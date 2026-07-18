# -*- coding: utf-8 -*-
"""
Điền form báo giá Excel + xuất PDF A4.

  python scripts/export-quote-xlsx.py --quote-id 1 --out data/exports/BG.xlsx
  python scripts/export-quote-xlsx.py --quote-id 1 --out data/exports/BG.xlsx --pdf
"""
from __future__ import annotations

import argparse
import re
import shutil
import sqlite3
import sys
import tempfile
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.worksheet.page import PageMargins
from PIL import Image as PILImage

CRM = Path(r"C:\Users\dankiet\Documents\CRM")
DB = CRM / "data" / "crm.db"
TEMPLATE = CRM / "templates" / "Bao_Gia_Form_V2.xlsx"
PUBLIC = CRM / "public"

ITEM_START = 19
ITEM_END_SAMPLE = 35
TOTAL_LABEL_ROW = 36


def parse_date_display(s: str) -> str:
    if not s:
        return datetime.now().strftime("%d/%m/%Y")
    m = re.match(r"(\d{4})-(\d{2})-(\d{2})", s)
    if m:
        return f"{m.group(3)}/{m.group(2)}/{m.group(1)}"
    m = re.match(r"(\d{2})/(\d{2})/(\d{4})", s)
    if m:
        return s[:10]
    m = re.match(r"(\d{4})-(\d{2})-(\d{2})[ T]", s)
    if m:
        return f"{m.group(3)}/{m.group(2)}/{m.group(1)}"
    return datetime.now().strftime("%d/%m/%Y")


def clear_row_values(ws, row: int, cols=range(2, 15)) -> None:
    for c in cols:
        cell = ws.cell(row, c)
        try:
            cell.value = None
        except AttributeError:
            continue


def resolve_image_path(image_path: str | None) -> Path | None:
    if not image_path:
        return None
    p = image_path.strip()
    if not p:
        return None
    if p.startswith("/"):
        local = PUBLIC / p.lstrip("/").replace("/", "\\")
    else:
        local = Path(p)
    if local.exists() and local.is_file():
        return local
    return None


def make_thumb(src: Path, dest: Path, max_px: int = 160) -> Path | None:
    """Resize for Excel embed (keep aspect)."""
    try:
        with PILImage.open(src) as im:
            im = im.convert("RGB") if im.mode not in ("RGB", "RGBA") else im
            if im.mode == "RGBA":
                bg = PILImage.new("RGB", im.size, (255, 255, 255))
                bg.paste(im, mask=im.split()[-1])
                im = bg
            im.thumbnail((max_px, max_px), PILImage.Resampling.LANCZOS)
            dest.parent.mkdir(parents=True, exist_ok=True)
            im.save(dest, format="JPEG", quality=85)
            return dest
    except Exception:
        return None


def export_pdf_a4(xlsx_path: Path, pdf_path: Path) -> None:
    """Excel COM → PDF, khổ A4 landscape fit-to-page."""
    try:
        import win32com.client  # type: ignore
    except ImportError as e:
        raise RuntimeError("Cần pywin32 + Microsoft Excel để xuất PDF") from e

    excel = win32com.client.DispatchEx("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    wb = None
    try:
        wb = excel.Workbooks.Open(str(xlsx_path.resolve()))
        ws = wb.Worksheets(1)
        # A4 = 9, Landscape = 2 (xlLandscape)
        ws.PageSetup.PaperSize = 9
        ws.PageSetup.Orientation = 2
        # Fit width to 1 page; tall = False means unlimited height pages
        ws.PageSetup.Zoom = False
        try:
            ws.PageSetup.FitToPagesWide = 1
            ws.PageSetup.FitToPagesTall = False
        except Exception:
            try:
                ws.PageSetup.FitToPagesWide = 1
                ws.PageSetup.FitToPagesTall = 1
            except Exception:
                ws.PageSetup.Zoom = 70
        try:
            ws.PageSetup.LeftMargin = excel.InchesToPoints(0.25)
            ws.PageSetup.RightMargin = excel.InchesToPoints(0.25)
            ws.PageSetup.TopMargin = excel.InchesToPoints(0.3)
            ws.PageSetup.BottomMargin = excel.InchesToPoints(0.3)
        except Exception:
            pass
        # 0 = xlTypePDF
        pdf_path.parent.mkdir(parents=True, exist_ok=True)
        if pdf_path.exists():
            pdf_path.unlink()
        wb.ExportAsFixedFormat(0, str(pdf_path.resolve()))
    finally:
        if wb is not None:
            wb.Close(False)
        excel.Quit()


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--quote-id", type=int, required=True)
    ap.add_argument("--out", type=str, required=True)
    ap.add_argument("--pdf", action="store_true", help="Also export PDF A4")
    ap.add_argument("--pdf-only", action="store_true", help="Return PDF path as main out")
    args = ap.parse_args()

    if not TEMPLATE.exists():
        raise SystemExit(f"Missing template: {TEMPLATE}")
    if not DB.exists():
        raise SystemExit(f"Missing DB: {DB}")

    conn = sqlite3.connect(str(DB))
    conn.row_factory = sqlite3.Row

    q = conn.execute(
        """
        SELECT q.*, c.name AS customer_name, c.phone AS customer_phone,
               c.region AS customer_region, c.note AS customer_note,
               c.source AS customer_source
        FROM quotes q
        JOIN customers c ON c.id = q.customer_id
        WHERE q.id = ?
        """,
        (args.quote_id,),
    ).fetchone()
    if not q:
        raise SystemExit(f"Quote not found: {args.quote_id}")

    items = conn.execute(
        """
        SELECT qi.*,
          COALESCE(
            (SELECT path FROM product_images pi
             WHERE pi.product_id = qi.product_id
             ORDER BY pi.is_primary DESC, pi.sort_order ASC, pi.id ASC
             LIMIT 1),
            p.image_path
          ) AS product_image,
          p.material AS product_material,
          COALESCE(p.packing, '') AS packing,
          p.packing_m2 AS packing_m2,
          p.packing_pcs AS packing_pcs
        FROM quote_items qi
        LEFT JOIN products p ON p.id = qi.product_id
        WHERE qi.quote_id = ?
        ORDER BY qi.id
        """,
        (args.quote_id,),
    ).fetchall()
    conn.close()

    out_arg = Path(args.out)
    out_arg.parent.mkdir(parents=True, exist_ok=True)
    if args.pdf_only or out_arg.suffix.lower() == ".pdf":
        xlsx_path = out_arg.with_suffix(".xlsx")
        pdf_path = out_arg if out_arg.suffix.lower() == ".pdf" else out_arg.with_suffix(".pdf")
        want_pdf = True
    else:
        xlsx_path = out_arg if out_arg.suffix.lower() == ".xlsx" else out_arg.with_suffix(".xlsx")
        pdf_path = xlsx_path.with_suffix(".pdf")
        want_pdf = args.pdf or args.pdf_only

    shutil.copy2(TEMPLATE, xlsx_path)
    wb = load_workbook(xlsx_path)
    ws = wb["BAO GIA"]

    # Header
    ws["M4"] = q["code"]
    ws["M5"] = parse_date_display(q["created_at"] or "")
    ws["B8"] = f"Kính gửi: {q['customer_name'] or ''}".strip()
    cong_trinh = (q["customer_note"] or "").strip() or "—"
    ws["B9"] = f"Công trình: {cong_trinh}"
    ws["H9"] = f"Người liên lạc: {q['customer_name'] or ''}"
    dia_diem = (q["customer_region"] or "").strip()
    ws["B10"] = f"Địa điểm giao hàng: {dia_diem or '—'}"
    ws["B13"] = "Mr Kiệt - 0901 378 391"

    n = len(items)
    sample_count = ITEM_END_SAMPLE - ITEM_START + 1

    if n > sample_count:
        extra = n - sample_count
        ws.insert_rows(ITEM_END_SAMPLE + 1, amount=extra)
        for i in range(extra):
            r = ITEM_END_SAMPLE + 1 + i
            ws.row_dimensions[r].height = ws.row_dimensions[ITEM_START].height

    end_items = ITEM_START + max(n, sample_count) - 1
    for r in range(ITEM_START, end_items + 1):
        clear_row_values(ws, r)

    # Remove existing floating images in data area if any (template sample)
    # Keep header logo images: only clear images anchored at rows >= ITEM_START
    try:
        keep = []
        for img in list(ws._images):
            anchor = getattr(img, "anchor", None)
            row = None
            if hasattr(anchor, "_from"):
                row = anchor._from.row + 1  # 0-indexed
            if row is not None and row >= ITEM_START:
                continue
            keep.append(img)
        ws._images = keep
    except Exception:
        pass

    thumb_dir = Path(tempfile.mkdtemp(prefix="quote_thumbs_"))

    for idx, it in enumerate(items):
        r = ITEM_START + idx
        ws.row_dimensions[r].height = max(ws.row_dimensions[r].height or 45, 48)
        ws.cell(r, 2).value = idx + 1
        ws.cell(r, 4).value = it["product_code"]
        ws.cell(r, 5).value = it["product_name"]
        size = (it["size"] or "").strip()
        if size and "mm" not in size.lower() and "cm" not in size.lower():
            size = f"{size}mm"
        ws.cell(r, 6).value = size
        mat = (it["product_material"] or "").strip()
        # bề mặt: suy từ tên nếu material trống
        name_l = (it["product_name"] or "").lower()
        if not mat:
            if "bóng" in name_l:
                mat = "Bóng"
            elif "mờ" in name_l:
                mat = "Mờ"
            else:
                mat = "—"
        ws.cell(r, 7).value = mat
        packing = (it["packing"] or "").strip() or "—"
        ws.cell(r, 8).value = packing
        qty = float(it["quantity_m2"] or 0)
        ws.cell(r, 9).value = qty
        ws.cell(r, 10).value = "m2"
        price = int(it["unit_price"] or 0)
        ws.cell(r, 11).value = price
        ws.cell(r, 12).value = f"=K{r}*I{r}"
        ws.cell(r, 13).value = "—"
        note = (it["area"] or "").strip()
        # ghi chú thêm số thùng nếu có packing_m2
        extra_note = []
        if note:
            extra_note.append(note)
        m2b = it["packing_m2"]
        if m2b and isinstance(m2b, (int, float)) and m2b > 0 and qty > 0:
            boxes = qty / float(m2b)
            boxes_s = f"{boxes:.1f}".rstrip("0").rstrip(".")
            extra_note.append(f"~{boxes_s} thùng")
        ws.cell(r, 14).value = " · ".join(extra_note) if extra_note else ""

        # Primary image
        img_path = resolve_image_path(it["product_image"])
        if img_path:
            thumb = make_thumb(img_path, thumb_dir / f"t{idx}.jpg")
            if thumb and thumb.exists():
                try:
                    xl_img = XLImage(str(thumb))
                    xl_img.width = 58
                    xl_img.height = 42
                    xl_img.anchor = f"C{r}"
                    ws.add_image(xl_img)
                except Exception:
                    pass

    last_item_row = ITEM_START + n - 1 if n else ITEM_START - 1
    if n > sample_count:
        total_row = last_item_row + 1
        vat_row = total_row + 1
        pay_row = total_row + 2
    else:
        total_row = TOTAL_LABEL_ROW
        vat_row = 37
        pay_row = 38
        for r in range(last_item_row + 1, ITEM_END_SAMPLE + 1):
            clear_row_values(ws, r)

    if n == 0:
        ws.cell(total_row, 8).value = "TỔNG CỘNG"
        ws.cell(total_row, 12).value = 0
        ws.cell(vat_row, 8).value = "VAT (8%)"
        ws.cell(vat_row, 12).value = 0
        ws.cell(pay_row, 8).value = "TỔNG THANH TOÁN"
        ws.cell(pay_row, 12).value = 0
    else:
        ws.cell(total_row, 8).value = "TỔNG CỘNG"
        ws.cell(total_row, 12).value = f"=SUM(L{ITEM_START}:L{last_item_row})"
        ws.cell(vat_row, 8).value = "VAT (8%)"
        ws.cell(vat_row, 12).value = f"=L{total_row}*0.08"
        ws.cell(pay_row, 8).value = "TỔNG THANH TOÁN"
        ws.cell(pay_row, 12).value = f"=L{total_row}+L{vat_row}"

    # Page setup for print (openpyxl — Excel COM re-applies fit for PDF)
    # Note: do NOT set page_setup.fitToPage (AttributeError on many openpyxl versions).
    try:
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr.fitToPage = True
    except Exception:
        pass
    try:
        ws.page_margins = PageMargins(
            left=0.25, right=0.25, top=0.3, bottom=0.3, header=0.1, footer=0.1
        )
        ws.print_options.horizontalCentered = True
    except Exception:
        pass

    notes = (q["notes"] or "").strip()
    if notes:
        try:
            prev = ws["B47"].value
            ws["B47"] = f"{prev}\nGhi chú: {notes}" if prev else f"Ghi chú: {notes}"
        except Exception:
            pass

    wb.save(xlsx_path)

    result_path = xlsx_path
    if want_pdf:
        try:
            export_pdf_a4(xlsx_path, pdf_path)
            result_path = pdf_path
            print(pdf_path)
        except Exception as e:
            print(f"PDF_ERROR: {e}", file=sys.stderr)
            print(xlsx_path)
            sys.exit(2)
    else:
        print(xlsx_path)

    # cleanup thumbs
    try:
        shutil.rmtree(thumb_dir, ignore_errors=True)
    except Exception:
        pass


if __name__ == "__main__":
    main()
