# -*- coding: utf-8 -*-
"""
Import tồn kho (m²) từ Stock.xlsx — map theo mã nội bộ HHDV.

  Stock.xlsx:
    row 0-1: title
    row 2: STT | Mã | Số lượng tồn
    data from row 3...

  Match: Stock.Mã  ==  products.internal_code  (case-insensitive)
  Đơn vị: m². Trùng mã → cộng dồn số lượng.

  python scripts/import-stock-from-xlsx.py
  python scripts/import-stock-from-xlsx.py --xlsx Stock.xlsx
"""
from __future__ import annotations

import argparse
import sqlite3
from pathlib import Path

from openpyxl import load_workbook

CRM = Path(r"C:\Users\dankiet\Documents\CRM")
DB = CRM / "data" / "crm.db"
DEFAULT_XLSX = CRM / "Stock.xlsx"


def norm_code(v) -> str:
    if v is None:
        return ""
    if isinstance(v, float):
        if v == int(v):
            return str(int(v))
        return str(v).strip()
    if isinstance(v, int):
        return str(v)
    s = str(v).strip()
    if s.endswith(".0") and s[:-2].replace("-", "").isdigit():
        return s[:-2]
    return s


def norm_qty(v) -> float:
    if v is None or v == "":
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return 0.0


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", type=str, default=str(DEFAULT_XLSX))
    ap.add_argument("--db", type=str, default=str(DB))
    ap.add_argument(
        "--clear-unmatched",
        action="store_true",
        help="Set stock_m2 = NULL for products not found in stock file",
    )
    args = ap.parse_args()

    xlsx = Path(args.xlsx)
    db_path = Path(args.db)
    if not xlsx.exists():
        raise SystemExit(f"Missing stock file: {xlsx}")
    if not db_path.exists():
        raise SystemExit(f"Missing DB: {db_path}")

    wb = load_workbook(str(xlsx), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    # upper(code) -> sum qty m2
    stock: dict[str, float] = {}
    rows_read = 0
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 3:
            continue  # title + header
        ma = norm_code(row[1] if len(row) > 1 else None)
        if not ma:
            continue
        qty = norm_qty(row[2] if len(row) > 2 else None)
        key = ma.upper()
        stock[key] = stock.get(key, 0.0) + qty
        rows_read += 1
    wb.close()

    conn = sqlite3.connect(str(db_path), timeout=60)
    conn.execute("PRAGMA busy_timeout=60000")
    cols = {r[1] for r in conn.execute("PRAGMA table_info(products)")}
    if "stock_m2" not in cols:
        conn.execute("ALTER TABLE products ADD COLUMN stock_m2 REAL")
        conn.commit()
    if "internal_code" not in cols:
        raise SystemExit(
            "Thiếu cột internal_code — chạy import-internal-code-from-mapping.py trước"
        )

    products = conn.execute(
        "SELECT id, code, internal_code, stock_m2 FROM products"
    ).fetchall()

    # internal_code upper -> list of product ids (rare dups)
    by_ic: dict[str, list[int]] = {}
    for pid, code, ic, _ in products:
        if not ic:
            continue
        by_ic.setdefault(str(ic).upper(), []).append(pid)

    updated = 0
    matched_codes = 0
    no_product = 0
    total_m2_applied = 0.0
    samples: list[tuple[str, float]] = []

    cur = conn.cursor()
    for key, qty in stock.items():
        pids = by_ic.get(key)
        if not pids:
            no_product += 1
            continue
        matched_codes += 1
        # round to 3 decimals (file uses thousandths)
        qty_r = round(qty, 3)
        for pid in pids:
            cur.execute(
                "UPDATE products SET stock_m2 = ? WHERE id = ?",
                (qty_r, pid),
            )
            updated += 1
            total_m2_applied += qty_r
            if len(samples) < 8:
                samples.append((key, qty_r))

    if args.clear_unmatched:
        matched_ids = set()
        for key in stock:
            for pid in by_ic.get(key, []):
                matched_ids.add(pid)
        for pid, code, ic, _ in products:
            if pid not in matched_ids:
                cur.execute(
                    "UPDATE products SET stock_m2 = NULL WHERE id = ?", (pid,)
                )

    conn.commit()

    with_stock = conn.execute(
        "SELECT COUNT(1) FROM products WHERE stock_m2 IS NOT NULL"
    ).fetchone()[0]
    positive = conn.execute(
        "SELECT COUNT(1) FROM products WHERE stock_m2 IS NOT NULL AND stock_m2 > 0"
    ).fetchone()[0]
    sum_stock = conn.execute(
        "SELECT COALESCE(SUM(stock_m2), 0) FROM products WHERE stock_m2 IS NOT NULL"
    ).fetchone()[0]
    total = conn.execute("SELECT COUNT(1) FROM products").fetchone()[0]
    conn.close()

    print(f"Stock rows read: {rows_read} | unique codes: {len(stock)}")
    print(f"Matched internal_code: {matched_codes} | no product: {no_product}")
    print(f"Products updated: {updated}")
    print(f"Products with stock_m2 set: {with_stock}/{total} ( >0: {positive} )")
    print(f"Sum stock_m2 in DB: {round(float(sum_stock), 2)} m2")
    if samples:
        print("Samples (mã nội bộ → m2):")
        for ma, q in samples:
            print(f"  {ma}: {q}")


if __name__ == "__main__":
    main()
