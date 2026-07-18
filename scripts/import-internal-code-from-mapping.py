# -*- coding: utf-8 -*-
"""
Import mã nội bộ (Mã HHDV) từ file Mapping_HHDV_BaoGia_NXT.

  Col A: Mã số (Báo giá)  → products.code
  Col B: Mã HHDV          → products.internal_code

  python scripts/import-internal-code-from-mapping.py
  python scripts/import-internal-code-from-mapping.py --xlsx "C:\\path\\to\\Mapping.xlsx"
"""
from __future__ import annotations

import argparse
import sqlite3
from pathlib import Path

from openpyxl import load_workbook

CRM = Path(r"C:\Users\dankiet\Documents\CRM")
DB = CRM / "data" / "crm.db"
DEFAULT_XLSX = Path(
    r"C:\Users\dankiet\Downloads\Mapping_HHDV_BaoGia_NXT_BanChay.xlsx"
)


def norm_code(v) -> str:
    if v is None:
        return ""
    if isinstance(v, float):
        if v == int(v):
            return str(int(v))
        return str(v).strip()
    if isinstance(v, int):
        return str(v)
    s = str(v).strip()
    # Excel sometimes gives "20140.0"
    if s.endswith(".0") and s[:-2].replace("-", "").isdigit():
        return s[:-2]
    return s


def hhdv_rank(h: str) -> tuple:
    """Lower = better. Prefer base HHDV over regional (-HN) / suffix variants."""
    u = h.upper().replace(" ", "")
    regional = 1 if u.endswith("-HN") or "-HN" in u else 0
    # trailing -1, -2 variants (e.g. M7532P2-1) slightly lower priority than base
    import re

    numbered = 1 if re.search(r"-\d+$", u) else 0
    return (regional, numbered, len(u))


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", type=str, default=str(DEFAULT_XLSX))
    ap.add_argument("--db", type=str, default=str(DB))
    args = ap.parse_args()

    xlsx = Path(args.xlsx)
    db_path = Path(args.db)
    if not xlsx.exists():
        raise SystemExit(f"Missing mapping file: {xlsx}")
    if not db_path.exists():
        raise SystemExit(f"Missing DB: {db_path}")

    wb = load_workbook(str(xlsx), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    # quote_code (upper) -> hhdv (best of duplicates)
    mapping: dict[str, str] = {}
    skipped_no_quote = 0
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        quote = norm_code(row[0] if len(row) > 0 else None)
        hhdv = norm_code(row[1] if len(row) > 1 else None)
        if not hhdv:
            continue
        if not quote:
            skipped_no_quote += 1
            continue
        key = quote.upper()
        prev = mapping.get(key)
        if prev is None or hhdv_rank(hhdv) < hhdv_rank(prev):
            mapping[key] = hhdv
    wb.close()

    conn = sqlite3.connect(str(db_path))
    # ensure column
    cols = {r[1] for r in conn.execute("PRAGMA table_info(products)")}
    if "internal_code" not in cols:
        conn.execute(
            "ALTER TABLE products ADD COLUMN internal_code TEXT NOT NULL DEFAULT ''"
        )
        conn.commit()

    products = conn.execute("SELECT id, code, internal_code FROM products").fetchall()
    by_code = {str(code).upper(): (pid, code, ic or "") for pid, code, ic in products}

    updated = 0
    same = 0
    missing = 0
    samples: list[tuple[str, str, str]] = []

    cur = conn.cursor()
    for key, hhdv in mapping.items():
        row = by_code.get(key)
        if not row:
            missing += 1
            continue
        pid, code, old = row
        if old == hhdv:
            same += 1
            continue
        cur.execute(
            "UPDATE products SET internal_code = ? WHERE id = ?",
            (hhdv, pid),
        )
        updated += 1
        if len(samples) < 8:
            samples.append((code, old or "(trống)", hhdv))

    conn.commit()

    with_code = conn.execute(
        "SELECT COUNT(*) FROM products WHERE internal_code IS NOT NULL AND internal_code != ''"
    ).fetchone()[0]
    total = conn.execute("SELECT COUNT(*) FROM products").fetchone()[0]
    conn.close()

    print(f"Mapping pairs (có mã báo giá): {len(mapping)}")
    print(f"Rows Excel không có mã báo giá (bỏ): {skipped_no_quote}")
    print(f"Updated: {updated} | already same: {same} | no product match: {missing}")
    print(f"Products with internal_code: {with_code}/{total}")
    if samples:
        print("Samples:")
        for code, old, new in samples:
            print(f"  {code}: {old} → {new}")


if __name__ == "__main__":
    main()
